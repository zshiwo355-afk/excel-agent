from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.agent.graph import execute_graph, plan_graph
from app.agent.state import AgentState


BASE_DIR = Path(__file__).resolve().parent
TMP_DIR = BASE_DIR / "storage" / "tmp_split_test"


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售明细"
    sheet.append(["公司", "产品", "金额"])
    sheet.append(["A公司", "产品1", 100])
    sheet.append(["B公司", "产品2", 200])
    sheet.append(["A公司", "产品3", 300])
    workbook.save(path)


def run_split_test() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    file_path = TMP_DIR / "companies.xlsx"
    write_workbook(file_path)

    state = AgentState(
        task_id="split-test",
        message="帮我把里面的公司拆出来，一个公司一个表",
        uploaded_file_path=str(file_path),
        uploaded_file_paths=[str(file_path)],
        uploaded_files=[
            {
                "file_id": "file_1",
                "file_name": "companies.xlsx",
                "file_path": str(file_path),
                "size": file_path.stat().st_size,
            }
        ],
        logs=["Task created."],
        status="planning",
    )

    planned = AgentState.model_validate(plan_graph.invoke(state))
    assert planned.excel_plan is not None
    assert planned.excel_plan["action"] == "modify_workbook"
    assert planned.excel_plan["sheets"][0]["operation"] == "split_sheet_by_column"
    assert planned.excel_plan["sheets"][0]["split"]["column"] == "公司"

    executed = AgentState.model_validate(execute_graph.invoke(planned.model_copy(update={"status": "running"})))
    assert executed.status == "completed"
    assert executed.output_file_path

    workbook = load_workbook(executed.output_file_path)
    assert "A公司" in workbook.sheetnames
    assert "B公司" in workbook.sheetnames
    assert workbook["A公司"].max_row - 1 == 2
    assert workbook["B公司"].max_row - 1 == 1
    assert (workbook["A公司"].max_row - 1) + (workbook["B公司"].max_row - 1) == 3


if __name__ == "__main__":
    run_split_test()
    print("split workbook test passed")
