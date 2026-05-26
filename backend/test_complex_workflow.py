from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.agent.graph import execute_graph, plan_graph
from app.agent.state import AgentState


BASE_DIR = Path(__file__).resolve().parent
TMP_DIR = BASE_DIR / "storage" / "tmp_complex_test"


def write_workbook(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def run_complex_workflow() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    file1 = TMP_DIR / "file1.xlsx"
    file2 = TMP_DIR / "file2.xlsx"

    write_workbook(
        file1,
        ["门店名称", "地区", "产品", "金额"],
        [["A店", "华东", "产品1", 100], ["B店", "华北", "产品2", 200]],
    )
    write_workbook(
        file2,
        ["药店名称", "区域", "商品名称", "销售额"],
        [["C店", "华东", "产品1", 300], ["A店", "华东", "产品1", 100]],
    )

    state = AgentState(
        task_id="complex-test",
        message="把这两个 Excel 合并成一个总表，字段尽量对齐，去掉重复数据，按地区汇总销售额，并生成一个汇总 sheet。",
        uploaded_file_path=str(file1),
        uploaded_file_paths=[str(file1), str(file2)],
        uploaded_files=[
            {"file_id": "file_1", "file_name": "file1.xlsx", "file_path": str(file1), "size": file1.stat().st_size},
            {"file_id": "file_2", "file_name": "file2.xlsx", "file_path": str(file2), "size": file2.stat().st_size},
        ],
        logs=["Task created."],
        status="planning",
    )
    planned = AgentState.model_validate(plan_graph.invoke(state))
    assert planned.task_mode == "complex"
    assert planned.task_plan is not None
    assert planned.status == "waiting_confirm"

    running_state = AgentState.model_validate(
        execute_graph.invoke(
            planned.model_copy(update={"status": "running"})
        )
    )
    assert running_state.status == "waiting_step_confirm"
    assert running_state.pending_step_id is not None

    resumed = running_state.model_copy(
        update={
            "status": "running",
            "confirmed_step_ids": [running_state.pending_step_id],
        }
    )
    finished = AgentState.model_validate(execute_graph.invoke(resumed))
    assert finished.status == "completed"
    assert finished.output_file_path

    workbook = load_workbook(finished.output_file_path)
    assert "合并明细" in workbook.sheetnames
    assert "地区汇总" in workbook.sheetnames
    detail_sheet = workbook["合并明细"]
    headers = [detail_sheet.cell(row=1, column=idx).value for idx in range(1, detail_sheet.max_column + 1)]
    assert "来源文件" in headers
    assert "来源Sheet" in headers
    assert detail_sheet.max_row - 1 == 3

    summary_sheet = workbook["地区汇总"]
    summary = {
        summary_sheet.cell(row=row_idx, column=1).value: summary_sheet.cell(row=row_idx, column=2).value
        for row_idx in range(2, summary_sheet.max_row + 1)
    }
    assert summary["华东"] == 400
    assert summary["华北"] == 200


if __name__ == "__main__":
    run_complex_workflow()
    print("complex workflow test passed")
