from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.agent.graph import execute_graph, plan_graph
from app.agent.state import AgentState


BASE_DIR = Path(__file__).resolve().parent
TMP_DIR = BASE_DIR / "storage" / "tmp_simple_format_test"


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["姓名", "分数"])
    sheet.append(["张三", 90])
    sheet.append(["李四", 85])
    workbook.save(path)


def run_simple_format_steps_test() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    file_path = TMP_DIR / "input.xlsx"
    write_workbook(file_path)

    state = AgentState(
        task_id="simple-format-test",
        message="帮我格式化表格，冻结表头，开启筛选，自动列宽，表头加粗",
        uploaded_file_path=str(file_path),
        uploaded_file_paths=[str(file_path)],
        uploaded_files=[
            {
                "file_id": "file_1",
                "file_name": "input.xlsx",
                "file_path": str(file_path),
                "size": file_path.stat().st_size,
            }
        ],
        logs=["Task created."],
        status="planning",
    )

    planned = AgentState.model_validate(plan_graph.invoke(state))
    assert planned.task_mode == "simple"
    assert planned.status == "waiting_confirm"

    executed = AgentState.model_validate(execute_graph.invoke(planned.model_copy(update={"status": "running"})))
    assert executed.status == "completed"
    assert executed.output_file_path

    titles = [step.title for step in executed.execution_steps]
    assert "冻结表头" in titles
    assert "开启筛选" in titles
    assert "自动调整列宽" in titles
    assert "表头加粗" in titles
    assert "生成结果文件" in titles
    assert "校验结果文件" in titles
    assert all(step.status == "completed" for step in executed.execution_steps)

    workbook = load_workbook(executed.output_file_path)
    sheet = workbook["Sheet1"]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None
    assert sheet["A1"].font.bold is True
    assert sheet.column_dimensions["A"].width is not None


if __name__ == "__main__":
    run_simple_format_steps_test()
    print("simple format steps test passed")
