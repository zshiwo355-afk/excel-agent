from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.agent.graph import execute_graph, plan_graph
from app.agent.state import AgentState


BASE_DIR = Path(__file__).resolve().parent
TMP_DIR = BASE_DIR / "storage" / "tmp_template_test"


def write_data_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据表"
    sheet.append(["商品", "价格", "地区"])
    sheet.append(["黄金A", 630, "华东"])
    sheet.append(["黄金B", 680, "华北"])
    sheet.append(["黄金C", 650, "华南"])
    workbook.save(path)


def write_template_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "模板表"
    sheet.append(["地区", "商品", "价格"])
    sheet["A1"].font = Font(bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    sheet["B1"].font = Font(bold=True, color="FFFFFF")
    sheet["B1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    sheet["C1"].font = Font(bold=True, color="FFFFFF")
    sheet["C1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 14
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="EAF3FF")
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="EAF3FF")
    sheet["C2"].fill = PatternFill(fill_type="solid", fgColor="EAF3FF")
    workbook.save(path)


def run_template_apply_test() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    data_file = TMP_DIR / "data.xlsx"
    template_file = TMP_DIR / "template.xlsx"
    write_data_workbook(data_file)
    write_template_workbook(template_file)

    state = AgentState(
        task_id="template-test",
        message="把表A里面的数据按照表B里面的格式进行排序，价格从高到低",
        uploaded_file_path=str(data_file),
        uploaded_file_paths=[str(data_file), str(template_file)],
        uploaded_files=[
            {
                "file_id": "file_1",
                "file_name": "data.xlsx",
                "file_path": str(data_file),
                "size": data_file.stat().st_size,
            },
            {
                "file_id": "file_2",
                "file_name": "template.xlsx",
                "file_path": str(template_file),
                "size": template_file.stat().st_size,
            },
        ],
        logs=["Task created."],
        status="planning",
    )

    planned = AgentState.model_validate(plan_graph.invoke(state))
    assert planned.excel_plan is not None
    assert planned.excel_plan["sheets"][0]["operation"] == "apply_template_sheet"

    executed = AgentState.model_validate(execute_graph.invoke(planned.model_copy(update={"status": "running"})))
    assert executed.status == "completed"
    assert executed.output_file_path

    workbook = load_workbook(executed.output_file_path)
    sheet = workbook["模板表"]
    assert [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value] == ["地区", "商品", "价格"]
    assert sheet["A2"].value == "华北"
    assert sheet["B2"].value == "黄金B"
    assert sheet["C2"].value == 680
    assert sheet["A3"].value == "华南"
    assert sheet["C4"].value == 630
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].fill.fgColor.rgb in {"001F4E78", "1F4E78"}
    assert sheet.column_dimensions["A"].width == 18


if __name__ == "__main__":
    run_template_apply_test()
    print("apply template sheet test passed")
