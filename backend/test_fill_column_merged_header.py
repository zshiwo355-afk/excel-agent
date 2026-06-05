from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.schemas.task import TaskDetail
from app.services.task_service import TaskService


BASE_DIR = Path(__file__).resolve().parent
TMP_DIR = BASE_DIR / "storage" / "tmp_fill_column_merged_header_test"


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "今日金价行情"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "黄金价格行情"
    sheet.append(["价格参考", "品类"])
    sheet.append([10, "A"])
    sheet.append([5, "B"])
    sheet.append([8, "C"])
    workbook.save(path)


def run_fill_column_merged_header_test() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    file_path = TMP_DIR / "input.xlsx"
    write_workbook(file_path)

    context = {
        "file_id": "file_1",
        "file_path": str(file_path),
        "file_name": "input.xlsx",
        "sheet_names": ["今日金价行情"],
        "sheets": [
            {
                "name": "今日金价行情",
                "header_row": 2,
                "data_start_row": 3,
                "headers": ["价格参考", "品类"],
                "merged_cells": ["A1:B1"],
            }
        ],
    }

    service = TaskService()
    task_id = "fill-column-merged-header-test"
    task = TaskDetail(
        task_id=task_id,
        message="把价格从低到高排序，然后插入日期，日期就是今天",
        status="running",
        auto_execute=True,
        uploaded_file_path=str(file_path),
        uploaded_file_paths=[str(file_path)],
        uploaded_files=[
            {
                "file_id": "file_1",
                "file_name": "input.xlsx",
                "file_path": str(file_path),
                "size": file_path.stat().st_size,
            }
        ],
        workbook_context=context,
        workbook_contexts=[context],
        excel_plan={
            "action": "modify_workbook",
            "workbook_name": "排序并填日期.xlsx",
            "sheets": [
                {
                    "operation": "sort_rows",
                    "name": "今日金价行情",
                    "source_sheet": "今日金价行情",
                    "header_row": 2,
                    "data_start_row": 3,
                    "sort": {
                        "column": "价格参考",
                        "order": "asc",
                        "numeric": True,
                    },
                },
                {
                    "operation": "fill_column_with_value",
                    "name": "今日金价行情",
                    "source_sheet": "今日金价行情",
                    "header_row": 2,
                    "data_start_row": 3,
                    "fill": {
                        "column_name": "日期",
                        "value_mode": "today_date",
                        "create_if_missing": True,
                        "overwrite_existing": True,
                    },
                },
            ],
        },
        logs=["Task created."],
        created_at="2026-01-01T00:00:00+00:00",
        updated_at="2026-01-01T00:00:00+00:00",
    )
    service.save_task(task)

    service.run_task_execution(task_id)
    executed = service.get_task(task_id)
    assert executed.status == "completed"
    assert executed.output_file_path

    workbook = load_workbook(executed.output_file_path)
    sheet = workbook["今日金价行情"]

    assert sheet["A1"].value == "黄金价格行情"
    assert sheet["A2"].value == "价格参考"
    assert sheet["B2"].value == "品类"
    assert sheet["C2"].value == "日期"

    assert sheet["A3"].value == 5
    assert sheet["A4"].value == 8
    assert sheet["A5"].value == 10
    assert sheet["C3"].value.date() == date.today()


if __name__ == "__main__":
    run_fill_column_merged_header_test()
    print("fill column merged header test passed")
