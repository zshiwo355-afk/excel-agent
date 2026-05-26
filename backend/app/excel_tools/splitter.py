from pathlib import Path

from openpyxl import Workbook

from app.excel_tools.formatter import apply_style_options
from app.excel_tools.reader import load_workbook_safe
from app.schemas.excel_plan import ExcelPlan


INVALID_SHEET_CHARS = '\\/?*[]:'


def _sanitize_sheet_name(value: str, used_names: set[str]) -> str:
    cleaned = "".join("_" if char in INVALID_SHEET_CHARS else char for char in str(value).strip())
    cleaned = cleaned or "未命名"
    cleaned = cleaned[:31] or "未命名"
    candidate = cleaned
    suffix = 2
    while candidate in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{cleaned[: max(0, 31 - len(suffix_text))]}{suffix_text}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def _resolve_column_index(sheet, header_row: int, column_name: str) -> int:
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=col_idx).value
        if value is not None and str(value).strip() == column_name:
            return col_idx
    raise ValueError(f"Split column '{column_name}' not found.")


def split_workbook_by_column(
    plan: ExcelPlan,
    source_file_path: str | Path,
    output_path: str | Path,
) -> dict[str, object]:
    split_sheet_plan = next(
        (sheet_plan for sheet_plan in plan.sheets if sheet_plan.operation == "split_sheet_by_column"),
        None,
    )
    if split_sheet_plan is None or split_sheet_plan.split is None or not split_sheet_plan.source_sheet:
        raise ValueError("split_sheet_by_column plan is missing required configuration.")

    source_workbook = load_workbook_safe(source_file_path, data_only=False)
    if split_sheet_plan.source_sheet not in source_workbook.sheetnames:
        raise ValueError(f"Source sheet '{split_sheet_plan.source_sheet}' not found.")
    source_sheet = source_workbook[split_sheet_plan.source_sheet]

    header_row = split_sheet_plan.header_row or 1
    data_start_row = split_sheet_plan.data_start_row or (header_row + 1)
    column_index = _resolve_column_index(source_sheet, header_row, split_sheet_plan.split.column)
    headers = [
        source_sheet.cell(row=header_row, column=col_idx).value
        for col_idx in range(1, source_sheet.max_column + 1)
    ]

    grouped_rows: dict[str, list[list[object]]] = {}
    for row_idx in range(data_start_row, source_sheet.max_row + 1):
        row_values = [
            source_sheet.cell(row=row_idx, column=col_idx).value
            for col_idx in range(1, source_sheet.max_column + 1)
        ]
        if all(value in (None, "") for value in row_values):
            continue
        group_value = row_values[column_index - 1]
        key = "未命名" if group_value in (None, "") else str(group_value).strip()
        grouped_rows.setdefault(key, []).append(row_values)

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()

    if split_sheet_plan.split.include_source_sheet:
        original_sheet = workbook.create_sheet(title=_sanitize_sheet_name(split_sheet_plan.source_sheet, used_names))
        original_sheet.append(headers)
        for row_idx in range(data_start_row, source_sheet.max_row + 1):
            row_values = [
                source_sheet.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, source_sheet.max_column + 1)
            ]
            if all(value in (None, "") for value in row_values):
                continue
            original_sheet.append(row_values)
        apply_style_options(original_sheet, split_sheet_plan.style)

    created_sheet_names: list[str] = []
    for group_value, rows in grouped_rows.items():
        sheet_name = (
            _sanitize_sheet_name(group_value, used_names)
            if split_sheet_plan.split.sanitize_sheet_name
            else str(group_value)
        )
        split_sheet = workbook.create_sheet(title=sheet_name)
        split_sheet.append(headers)
        for row_values in rows:
            split_sheet.append(row_values)
        apply_style_options(split_sheet, split_sheet_plan.style)
        created_sheet_names.append(sheet_name)

    workbook.save(output_path)
    return {
        "split_column": split_sheet_plan.split.column,
        "created_sheet_names": created_sheet_names,
        "total_split_sheets": len(created_sheet_names),
        "total_rows": sum(len(rows) for rows in grouped_rows.values()),
    }
