from pathlib import Path
from typing import Any

from openpyxl.cell.cell import Cell

from app.excel_tools.reader import load_workbook_safe
from app.utils.jsonable import to_jsonable


def _cell_value(cell: Cell):
    return cell.value


def _row_values(sheet, row_idx: int) -> list:
    return [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]


def _non_empty_values(values: list) -> list:
    return [value for value in values if value not in (None, "")]


def _detect_header_row(sheet) -> tuple[int, list[str], list[dict[str, object]]]:
    candidate_headers: list[dict[str, object]] = []
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(sheet.max_row, 10) + 1):
        values = _row_values(sheet, row_idx)
        non_empty = _non_empty_values(values)
        if not non_empty:
            continue

        merged_hits = [
            str(range_ref)
            for range_ref in sheet.merged_cells.ranges
            if range_ref.min_row <= row_idx <= range_ref.max_row
        ]
        if len(non_empty) == 1 and merged_hits:
            continue

        score = len(non_empty)
        candidate = {
            "row_index": row_idx,
            "non_empty_count": score,
            "values": [str(value).strip() for value in non_empty],
        }
        candidate_headers.append(candidate)

        if score > best_score:
            best_score = score
            best_row = row_idx

    headers = [
        str(value).strip()
        for value in _row_values(sheet, best_row)
        if value not in (None, "")
    ]
    return best_row, headers, candidate_headers


def profile_workbook(
    file_path: str | Path,
    *,
    file_id: str | None = None,
    file_name: str | None = None,
) -> dict[str, Any]:
    workbook = load_workbook_safe(file_path, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        header_row, headers, candidate_headers = _detect_header_row(sheet)
        data_start_row = min(header_row + 1, sheet.max_row) if sheet.max_row >= header_row + 1 else header_row
        sample_rows = []
        for row in sheet.iter_rows(
            min_row=data_start_row,
            max_row=min(sheet.max_row, data_start_row + 4),
            values_only=True,
        ):
            sample_rows.append(list(row))

        has_formula = any(
            isinstance(cell.value, str) and cell.value.startswith("=")
            for row in sheet.iter_rows()
            for cell in row
        )

        sheets.append(
            {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "header_row": header_row,
                "data_start_row": data_start_row,
                "candidate_headers": candidate_headers,
                "headers": headers,
                "sample_rows": sample_rows,
                "has_formula": has_formula,
                "has_merged_cells": bool(sheet.merged_cells.ranges),
                "merged_cells_count": len(sheet.merged_cells.ranges),
                "merged_cells": [str(item) for item in sheet.merged_cells.ranges],
            }
        )

    return to_jsonable(
        {
        "file_id": file_id,
        "file_path": str(Path(file_path).resolve()),
        "file_name": file_name or Path(file_path).name,
        "sheet_names": workbook.sheetnames,
        "sheets": sheets,
        }
    )
