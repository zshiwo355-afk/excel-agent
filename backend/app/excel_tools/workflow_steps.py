import shutil
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from app.config import get_settings
from app.excel_tools.formatter import apply_style_options
from app.excel_tools.merger import merge_workbooks_by_plan
from app.excel_tools.reader import load_workbook_safe
from app.excel_tools.validator import validate_output_workbook
from app.schemas.excel_plan import ExcelPlan, StylePlan
from app.schemas.task_plan import StepPlan, TaskPlan

WORKFLOW_COLUMN_ALIASES = {
    "地区": ["地区", "区域"],
    "金额": ["金额", "销售额", "价格"],
    "门店名称": ["门店名称", "药店名称", "门店"],
    "产品": ["产品", "商品", "商品名称"],
}


def _artifact_path(task_id: str, artifact_name: str) -> Path:
    settings = get_settings()
    output_dir = settings.outputs_dir / task_id
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{artifact_name}.xlsx"


def _copy_workbook(src: str | Path, dst: str | Path) -> Path:
    shutil.copy2(src, dst)
    return Path(dst)


def _step_by_id(task_plan: TaskPlan, step_id: str) -> StepPlan | None:
    for step in task_plan.steps:
        if step.step_id == step_id:
            return step
    return None


def _input_path_for_step(state, task_plan: TaskPlan, step: StepPlan) -> Path | None:
    if not step.input_artifact:
        return None
    path = state.step_artifacts.get(step.input_artifact)
    if path:
        return Path(path)
    for prev_step in task_plan.steps:
        if prev_step.output_artifact == step.input_artifact:
            candidate = state.step_artifacts.get(prev_step.output_artifact or "")
            if candidate:
                return Path(candidate)
    return None


def _headers(sheet) -> list[str]:
    return [str(sheet.cell(row=1, column=col_idx).value).strip() for col_idx in range(1, sheet.max_column + 1)]


def _resolve_column_name(headers: list[str], requested: str) -> str:
    if requested in headers:
        return requested
    for canonical, aliases in WORKFLOW_COLUMN_ALIASES.items():
        if requested == canonical or requested in aliases:
            for alias in aliases:
                if alias in headers:
                    return alias
    for header in headers:
        if requested in header or header in requested:
            return header
    raise KeyError(requested)


def execute_workflow_step(state, task_plan: TaskPlan, step: StepPlan) -> dict[str, Any]:
    if step.step_type == "analyze_files":
        return {
            "artifact_path": None,
            "result": {"workbook_count": len(state.workbook_contexts)},
        }

    if step.step_type == "confirm_column_mapping":
        if step.step_id not in state.confirmed_step_ids:
            return {
                "wait_for_confirm": True,
                "result": {"column_mapping": step.params.get("column_mapping", {})},
            }
        return {
            "artifact_path": None,
            "result": {"confirmed": True, "column_mapping": step.params.get("column_mapping", {})},
        }

    if step.step_type == "merge_workbooks":
        output_path = _artifact_path(state.task_id, step.output_artifact or "raw_merged_workbook")
        merge_plan_payload = step.params.get("merge_plan", {})
        plan = ExcelPlan.model_validate(
            {
                "action": "merge_workbooks",
                "workbook_name": output_path.name,
                "sheets": [],
                "merge": merge_plan_payload,
                "style": {
                    "freeze_header": True,
                    "auto_filter": True,
                    "auto_width": True,
                    "header_bold": True,
                },
            }
        )
        result = merge_workbooks_by_plan(plan, state.workbook_contexts, output_path)
        return {"artifact_path": str(output_path), "result": result}

    input_path = _input_path_for_step(state, task_plan, step)
    if step.step_type == "deduplicate_rows":
        if input_path is None:
            raise ValueError("deduplicate_rows 缺少输入工件。")
        output_path = _artifact_path(state.task_id, step.output_artifact or "deduplicated_workbook")
        _copy_workbook(input_path, output_path)
        workbook = load_workbook_safe(output_path, data_only=False)
        sheet = workbook[step.params.get("sheet_name", "合并明细")]
        headers = _headers(sheet)
        records = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
        frame = pd.DataFrame(records, columns=headers)
        before = len(frame.index)
        dedupe_columns = step.params.get("dedupe_columns") or [
            header for header in headers if header not in {"来源文件", "来源Sheet"}
        ]
        frame = frame.drop_duplicates(subset=dedupe_columns)
        after = len(frame.index)
        workbook.remove(sheet)
        new_sheet = workbook.create_sheet(title=step.params.get("sheet_name", "合并明细"), index=0)
        new_sheet.append(headers)
        for row in frame.itertuples(index=False):
            new_sheet.append(list(row))
        workbook.save(output_path)
        return {"artifact_path": str(output_path), "result": {"before_rows": before, "after_rows": after}}

    if step.step_type == "clean_rows":
        if input_path is None:
            raise ValueError("clean_rows 缺少输入工件。")
        output_path = _artifact_path(state.task_id, step.output_artifact or "cleaned_workbook")
        _copy_workbook(input_path, output_path)
        workbook = load_workbook_safe(output_path, data_only=False)
        for sheet in workbook.worksheets:
            rows_to_delete: list[int] = []
            for row_idx in range(2, sheet.max_row + 1):
                values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                if all(value in (None, "") for value in values):
                    rows_to_delete.append(row_idx)
                else:
                    for col_idx, value in enumerate(values, start=1):
                        if isinstance(value, str):
                            sheet.cell(row=row_idx, column=col_idx, value=value.strip())
            for row_idx in reversed(rows_to_delete):
                sheet.delete_rows(row_idx, 1)
        workbook.save(output_path)
        return {"artifact_path": str(output_path), "result": {"cleaned": True}}

    if step.step_type == "create_summary_sheet":
        if input_path is None:
            raise ValueError("create_summary_sheet 缺少输入工件。")
        output_path = _artifact_path(state.task_id, step.output_artifact or "summary_workbook")
        _copy_workbook(input_path, output_path)
        workbook = load_workbook_safe(output_path, data_only=False)
        source_sheet_name = step.params.get("source_sheet", "合并明细")
        target_sheet_name = step.params.get("target_sheet", "汇总")
        source_sheet = workbook[source_sheet_name]
        data = list(source_sheet.values)
        headers = list(data[0])
        frame = pd.DataFrame(data[1:], columns=headers)
        group_by = [
            _resolve_column_name(list(frame.columns), column_name)
            for column_name in step.params.get("group_by", ["地区"])
        ]
        metric_column = _resolve_column_name(list(frame.columns), step.params.get("metric_column", "金额"))
        metric_name = step.params.get("metric_name", "汇总金额")
        if target_sheet_name in workbook.sheetnames:
            del workbook[target_sheet_name]
        summary_sheet = workbook.create_sheet(title=target_sheet_name)
        grouped = frame.groupby(group_by, dropna=False)[metric_column].sum().reset_index()
        grouped = grouped.rename(columns={metric_column: metric_name})
        summary_sheet.append(list(grouped.columns))
        for row in grouped.itertuples(index=False):
            summary_sheet.append(list(row))
        apply_style_options(
            summary_sheet,
            StylePlan(freeze_header=True, auto_filter=True, auto_width=True, header_bold=True),
        )
        workbook.save(output_path)
        return {"artifact_path": str(output_path), "result": {"summary_rows": len(grouped.index), "sheet_name": target_sheet_name}}

    if step.step_type == "sort_rows":
        if input_path is None:
            raise ValueError("sort_rows 缺少输入工件。")
        output_path = _artifact_path(state.task_id, step.output_artifact or "sorted_workbook")
        _copy_workbook(input_path, output_path)
        workbook = load_workbook_safe(output_path, data_only=False)
        sheet = workbook[step.params.get("sheet_name", "合并明细")]
        headers = _headers(sheet)
        data = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
        frame = pd.DataFrame(data, columns=headers)
        sort_column = step.params.get("sort_column") or headers[0]
        ascending = step.params.get("order", "asc") != "desc"
        frame = frame.sort_values(by=sort_column, ascending=ascending)
        workbook.remove(sheet)
        new_sheet = workbook.create_sheet(title=step.params.get("sheet_name", "合并明细"), index=0)
        new_sheet.append(headers)
        for row in frame.itertuples(index=False):
            new_sheet.append(list(row))
        workbook.save(output_path)
        return {"artifact_path": str(output_path), "result": {"sorted_by": sort_column}}

    if step.step_type == "format_sheet":
        if input_path is None:
            raise ValueError("format_sheet 缺少输入工件。")
        output_path = _artifact_path(state.task_id, step.output_artifact or "formatted_workbook")
        _copy_workbook(input_path, output_path)
        workbook = load_workbook_safe(output_path, data_only=False)
        style = StylePlan(
            freeze_header=True,
            auto_filter=True,
            auto_width=True,
            header_bold=True,
        )
        sheet_names = step.params.get("sheet_names") or workbook.sheetnames
        for sheet_name in sheet_names:
            if sheet_name in workbook.sheetnames:
                apply_style_options(workbook[sheet_name], style)
        workbook.save(output_path)
        return {"artifact_path": str(output_path), "result": {"formatted_sheets": sheet_names}}

    if step.step_type == "create_chart":
        if input_path is None:
            raise ValueError("create_chart 缺少输入工件。")
        output_path = _artifact_path(state.task_id, step.output_artifact or "charted_workbook")
        _copy_workbook(input_path, output_path)
        workbook = load_workbook_safe(output_path, data_only=False)
        sheet = workbook[step.params.get("source_sheet", workbook.sheetnames[0])]
        chart = BarChart()
        chart.title = step.params.get("chart_title", "数据图表")
        if sheet.max_row >= 2 and sheet.max_column >= 2:
            data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
            cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            sheet.add_chart(chart, "E2")
        workbook.save(output_path)
        return {"artifact_path": str(output_path), "result": {"chart_added": True, "sheet_name": sheet.title}}

    if step.step_type == "create_exception_sheet":
        if input_path is None:
            raise ValueError("create_exception_sheet 缺少输入工件。")
        output_path = _artifact_path(state.task_id, step.output_artifact or "exception_workbook")
        _copy_workbook(input_path, output_path)
        workbook = load_workbook_safe(output_path, data_only=False)
        source_sheet = workbook[step.params.get("source_sheet", "合并明细")]
        target_sheet_name = step.params.get("target_sheet", "异常数据")
        headers = _headers(source_sheet)
        exception_rows = []
        for row in source_sheet.iter_rows(min_row=2, values_only=True):
            values = list(row)
            if any(value in (None, "") for value in values[: min(3, len(values))]):
                exception_rows.append(values)
        if target_sheet_name in workbook.sheetnames:
            del workbook[target_sheet_name]
        target_sheet = workbook.create_sheet(title=target_sheet_name)
        target_sheet.append(headers)
        for row in exception_rows:
            target_sheet.append(row)
        workbook.save(output_path)
        return {"artifact_path": str(output_path), "result": {"exception_rows": len(exception_rows), "sheet_name": target_sheet_name}}

    if step.step_type == "validate_workbook":
        if input_path is None:
            raise ValueError("validate_workbook 缺少输入工件。")
        validation = validate_output_workbook(input_path)
        return {"artifact_path": str(input_path), "result": validation}

    if step.step_type == "export_workbook":
        if input_path is None:
            raise ValueError("export_workbook 缺少输入工件。")
        workbook_name = step.params.get("workbook_name", "complex_workflow_result.xlsx")
        output_path = _artifact_path(state.task_id, Path(workbook_name).stem)
        _copy_workbook(input_path, output_path)
        return {"artifact_path": str(output_path), "result": {"exported": True}}

    raise ValueError(f"Unsupported step type: {step.step_type}")


def validate_workflow_step(state, task_plan: TaskPlan, step: StepPlan) -> dict[str, Any]:
    result = state.last_step_result or {}
    artifact_path = state.last_step_artifact

    if step.step_type == "analyze_files":
        ok = result.get("workbook_count", 0) >= step.validation.get("workbook_count_min", 1)
        return {"ok": ok, "message": "analyze_files validation passed." if ok else "分析文件数量不足。"}

    if step.step_type == "confirm_column_mapping":
        ok = bool(result.get("column_mapping"))
        return {"ok": ok, "message": "column mapping confirmed." if ok else "字段映射为空。"}

    if step.step_type == "merge_workbooks":
        ok = result.get("merged_rows", 0) > 0
        return {"ok": ok, "message": "merge validation passed." if ok else "合并结果为空。"}

    if step.step_type == "deduplicate_rows":
        ok = result.get("after_rows", 0) <= result.get("before_rows", 0)
        return {"ok": ok, "message": "deduplicate validation passed." if ok else "去重后行数异常。"}

    if step.step_type == "create_summary_sheet":
        workbook = load_workbook_safe(artifact_path, data_only=False)
        sheet_name = step.validation.get("sheet_name", step.params.get("target_sheet"))
        ok = sheet_name in workbook.sheetnames
        return {"ok": ok, "message": "summary sheet validation passed." if ok else "汇总 sheet 不存在。"}

    if step.step_type == "create_chart":
        workbook = load_workbook_safe(artifact_path, data_only=False)
        sheet = workbook[step.params.get("source_sheet", workbook.sheetnames[0])]
        ok = len(sheet._charts) > 0 or sheet.max_row > 1
        return {"ok": ok, "message": "chart validation passed." if ok else "图表或图表数据区域不存在。"}

    if step.step_type == "create_exception_sheet":
        workbook = load_workbook_safe(artifact_path, data_only=False)
        sheet_name = step.validation.get("sheet_name", step.params.get("target_sheet"))
        ok = sheet_name in workbook.sheetnames
        return {"ok": ok, "message": "exception sheet validation passed." if ok else "异常 sheet 不存在。"}

    if step.step_type == "validate_workbook":
        ok = bool(result.get("ok"))
        return {"ok": ok, "message": result.get("message", "Workbook validation completed.")}

    if step.step_type == "export_workbook":
        ok = bool(artifact_path and Path(artifact_path).exists())
        return {"ok": ok, "message": "export validation passed." if ok else "导出文件不存在。"}

    ok = True
    return {"ok": ok, "message": f"{step.step_type} validation passed."}
