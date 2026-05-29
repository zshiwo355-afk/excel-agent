from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.excel_plan import ExcelPlan, SheetPlan


def _sheet_headers(sheet: Worksheet, header_row: int) -> list[str]:
    headers: list[str] = []
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=col_idx).value
        headers.append(str(value).strip() if value not in (None, "") else "")
    return headers


def _copy_cell_style(src_sheet: Worksheet, src_row: int, src_col: int, dst_sheet: Worksheet, dst_row: int, dst_col: int) -> None:
    src_cell = src_sheet.cell(row=src_row, column=src_col)
    dst_cell = dst_sheet.cell(row=dst_row, column=dst_col)
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    if src_cell.font:
        dst_cell.font = copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy(src_cell.fill)
    if src_cell.border:
        dst_cell.border = copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy(src_cell.alignment)
    if src_cell.protection:
        dst_cell.protection = copy(src_cell.protection)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format


def _clear_data_rows(sheet: Worksheet, data_start_row: int) -> None:
    if sheet.max_row >= data_start_row:
        sheet.delete_rows(data_start_row, sheet.max_row - data_start_row + 1)


def _coerce_sort_value(value, numeric: bool):
    if value is None:
        return float("-inf") if numeric else ""
    if not numeric:
        return str(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.replace("¥", "").replace("元", "").replace(",", "").strip()
        try:
            return float(normalized)
        except ValueError:
            return float("-inf")
    return float("-inf")


def _build_sorted_rows(source_sheet: Worksheet, sheet_plan: SheetPlan, template_headers: list[str], mapping: dict[str, str]) -> list[dict[str, object]]:
    header_row = sheet_plan.header_row or 1
    data_start_row = sheet_plan.data_start_row or (header_row + 1)
    source_headers = _sheet_headers(source_sheet, header_row)
    source_index = {
        header: idx
        for idx, header in enumerate(source_headers)
        if header
    }

    rows: list[dict[str, object]] = []
    for row_idx in range(data_start_row, source_sheet.max_row + 1):
        row_values = [source_sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, source_sheet.max_column + 1)]
        if all(value in (None, "") for value in row_values):
            continue
        mapped_row: dict[str, object] = {}
        for template_header in template_headers:
            source_header = mapping.get(template_header) or template_header
            if source_header in source_index:
                mapped_row[template_header] = row_values[source_index[source_header]]
            else:
                mapped_row[template_header] = None
        rows.append(mapped_row)

    if sheet_plan.sort:
        sort_header = sheet_plan.sort.column
        reverse = sheet_plan.sort.order == "desc"
        rows.sort(
            key=lambda item: _coerce_sort_value(
                item.get(sort_header) if sort_header in item else item.get(next((key for key, val in mapping.items() if val == sort_header), sort_header)),
                sheet_plan.sort.numeric,
            ),
            reverse=reverse,
        )
    return rows


def _snapshot_style_row(sheet: Worksheet, row_idx: int, max_columns: int) -> list[dict[str, object]]:
    snapshot: list[dict[str, object]] = []
    for col_idx in range(1, max_columns + 1):
        cell = sheet.cell(row=row_idx, column=col_idx)
        snapshot.append(
            {
                "style": copy(cell._style) if cell.has_style else None,
                "font": copy(cell.font) if cell.font else None,
                "fill": copy(cell.fill) if cell.fill else None,
                "border": copy(cell.border) if cell.border else None,
                "alignment": copy(cell.alignment) if cell.alignment else None,
                "protection": copy(cell.protection) if cell.protection else None,
                "number_format": cell.number_format,
            }
        )
    return snapshot


def _apply_style_snapshot(sheet: Worksheet, row_idx: int, col_idx: int, style_snapshot: dict[str, object]) -> None:
    cell = sheet.cell(row=row_idx, column=col_idx)
    if style_snapshot.get("style") is not None:
        cell._style = copy(style_snapshot["style"])
    if style_snapshot.get("font") is not None:
        cell.font = copy(style_snapshot["font"])
    if style_snapshot.get("fill") is not None:
        cell.fill = copy(style_snapshot["fill"])
    if style_snapshot.get("border") is not None:
        cell.border = copy(style_snapshot["border"])
    if style_snapshot.get("alignment") is not None:
        cell.alignment = copy(style_snapshot["alignment"])
    if style_snapshot.get("protection") is not None:
        cell.protection = copy(style_snapshot["protection"])
    if style_snapshot.get("number_format"):
        cell.number_format = style_snapshot["number_format"]


def apply_template_sheet_from_plan(
    plan: ExcelPlan,
    uploaded_files: list[dict[str, object]],
    output_path: str | Path,
) -> Path:
    template_plan = next((sheet_plan for sheet_plan in plan.sheets if sheet_plan.operation == "apply_template_sheet"), None)
    if template_plan is None or template_plan.template is None:
        raise ValueError("apply_template_sheet plan is missing template configuration.")

    file_index = {str(item.get("file_id")): item for item in uploaded_files}
    source_file = file_index.get(template_plan.template.source_file_id or "")
    template_file = file_index.get(template_plan.template.template_file_id)
    if not source_file or not template_file:
        raise ValueError("Source file or template file was not found in uploaded files.")

    source_path = Path(str(source_file.get("file_path")))
    template_path = Path(str(template_file.get("file_path")))
    source_workbook = load_workbook(source_path, data_only=False)
    template_workbook = load_workbook(template_path, data_only=False)

    source_sheet_name = template_plan.template.source_sheet or template_plan.source_sheet or source_workbook.sheetnames[0]
    template_sheet_name = template_plan.template.template_sheet
    if source_sheet_name not in source_workbook.sheetnames:
        raise ValueError(f"Source sheet '{source_sheet_name}' not found.")
    if template_sheet_name not in template_workbook.sheetnames:
        raise ValueError(f"Template sheet '{template_sheet_name}' not found.")

    source_sheet = source_workbook[source_sheet_name]
    template_sheet = template_workbook[template_sheet_name]
    output_sheet_name = template_plan.template.output_sheet_name or template_sheet_name
    if template_sheet.title != output_sheet_name:
        template_sheet.title = output_sheet_name
    output_sheet = template_sheet

    template_header_row = 1
    template_data_start_row = template_plan.template.data_start_row or 2
    template_headers = [header for header in _sheet_headers(template_sheet, template_header_row) if header]
    mapping = template_plan.template.column_mapping or {header: header for header in template_headers}
    row_style_snapshot = _snapshot_style_row(template_sheet, template_data_start_row, max(1, len(template_headers)))

    if template_plan.template.clear_existing_data_rows:
        _clear_data_rows(output_sheet, template_data_start_row)

    sorted_rows = _build_sorted_rows(source_sheet, template_plan, template_headers, mapping)

    for row_offset, row_data in enumerate(sorted_rows, start=template_data_start_row):
        for col_idx, template_header in enumerate(template_headers, start=1):
            _apply_style_snapshot(output_sheet, row_offset, col_idx, row_style_snapshot[col_idx - 1])
            output_sheet.cell(row=row_offset, column=col_idx, value=row_data.get(template_header))

    template_workbook.save(output_path)
    return Path(output_path)
