from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from app.excel_tools.formatter import apply_style_options
from app.schemas.excel_plan import ExcelPlan, FormulaPlan, SheetPlan


def _write_headers_and_rows(sheet, plan: SheetPlan) -> None:
    if plan.columns:
        sheet.append(plan.columns)

    if plan.rows:
        for row in plan.rows:
            if plan.columns:
                sheet.append([row.get(col) for col in plan.columns])
            else:
                sheet.append(list(row.values()))
    elif plan.columns and plan.sample_rows:
        for _ in range(plan.sample_rows):
            sheet.append(["" for _ in plan.columns])


def _apply_formulas(sheet, formulas: list[FormulaPlan] | None) -> None:
    if not formulas:
        return
    for formula in formulas:
        sheet[formula.target_cell] = formula.expression


def _create_summary_sheet(workbook: Workbook, plan: SheetPlan) -> None:
    sheet = workbook.create_sheet(title=plan.name)
    headers = plan.group_by or []
    metric_names = [metric.name for metric in (plan.metrics or [])]
    sheet.append(headers + metric_names)
    apply_style_options(sheet, plan.style)


def create_workbook_from_plan(plan: ExcelPlan, output_path: str | Path) -> Path:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_plan in plan.sheets:
        if sheet_plan.operation == "create_summary_sheet":
            _create_summary_sheet(workbook, sheet_plan)
            continue

        sheet = workbook.create_sheet(title=sheet_plan.name)
        _write_headers_and_rows(sheet, sheet_plan)
        _apply_formulas(sheet, sheet_plan.formulas)
        apply_style_options(sheet, sheet_plan.style)

    if not workbook.sheetnames:
        workbook.create_sheet(title="Sheet1")

    workbook.save(output_path)
    return Path(output_path)
