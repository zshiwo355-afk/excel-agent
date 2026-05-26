from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.excel_tools.formatter import apply_style_options
from app.schemas.excel_plan import StylePlan


def run_smoke_test(output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MergedDemo"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "合并表头"
    sheet.append(["列1", "列2", "列3", "列4"])
    sheet.append(["a", "b", "c", "d"])
    workbook.save(path)

    workbook = load_workbook(path)
    sheet = workbook["MergedDemo"]
    apply_style_options(
        sheet,
        StylePlan(
            freeze_header=True,
            auto_filter=True,
            auto_width=True,
            header_bold=True,
        ),
    )
    workbook.save(path)
    return path


if __name__ == "__main__":
    result = run_smoke_test("storage/outputs/merged_cells_smoke.xlsx")
    print(result)
