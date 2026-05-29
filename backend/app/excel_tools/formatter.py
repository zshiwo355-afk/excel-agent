from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.excel_plan import StylePlan


def _is_real_cell(cell) -> bool:
    return not isinstance(cell, MergedCell)


def apply_header_bold(sheet: Worksheet) -> None:
    if sheet.max_row < 1:
        return
    for cell in sheet[1]:
        if not _is_real_cell(cell):
            continue
        cell.font = Font(bold=True)


def apply_freeze_header(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"


def apply_auto_filter(sheet: Worksheet) -> None:
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions


def apply_auto_width(sheet: Worksheet) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        max_len = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if not _is_real_cell(cell) or cell.value is None:
                continue
            value = str(cell.value)
            max_len = max(max_len, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 40)


def apply_style_options(sheet: Worksheet, style: StylePlan | None) -> None:
    if not style:
        return

    if style.header_bold and sheet.max_row >= 1:
        apply_header_bold(sheet)

    if style.freeze_header:
        apply_freeze_header(sheet)

    if style.auto_filter and sheet.max_row >= 1 and sheet.max_column >= 1:
        apply_auto_filter(sheet)

    if style.auto_width:
        apply_auto_width(sheet)
