import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.cell.cell import MergedCell

from app.excel_tools.formatter import (
    apply_auto_filter,
    apply_auto_width,
    apply_freeze_header,
    apply_header_bold,
    apply_style_options,
)
from app.excel_tools.reader import load_workbook_safe
from app.schemas.excel_plan import ExcelPlan, MetricPlan, SheetPlan


def _sheet_headers(sheet, header_row: int = 1) -> list[str]:
    headers: list[str] = []
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if isinstance(cell, MergedCell) or cell.value is None:
            continue
        value = str(cell.value).strip()
        if value:
            headers.append(value)
    return headers


def _next_writable_header_column(sheet, header_row: int = 1) -> int:
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        if cell.value in (None, ""):
            return col_idx
    return sheet.max_column + 1


def _append_columns(sheet, columns: list[str], header_row: int = 1) -> None:
    existing_headers = _sheet_headers(sheet, header_row)
    next_column = _next_writable_header_column(sheet, header_row)
    for column in columns:
        if column in existing_headers:
            continue
        sheet.cell(row=header_row, column=next_column, value=column)
        existing_headers.append(column)
        next_column += 1


def _ensure_column(sheet, column_name: str, header_row: int = 1) -> int:
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if isinstance(cell, MergedCell) or cell.value is None:
            continue
        if str(cell.value).strip() == column_name:
            return col_idx

    next_column = _next_writable_header_column(sheet, header_row)
    sheet.cell(row=header_row, column=next_column, value=column_name)
    return next_column


def _create_summary_from_source(workbook, plan: SheetPlan) -> None:
    if not plan.source_sheet:
        raise ValueError(f"Summary sheet '{plan.name}' requires source_sheet.")
    if not plan.group_by:
        raise ValueError(f"Summary sheet '{plan.name}' requires group_by.")
    if not plan.metrics:
        raise ValueError(f"Summary sheet '{plan.name}' requires metrics.")

    source_sheet = workbook[plan.source_sheet]
    data = source_sheet.values
    headers = next(data, None)
    if not headers:
        raise ValueError(f"Source sheet '{plan.source_sheet}' has no header row.")

    frame = pd.DataFrame(data, columns=headers)
    agg_map: dict[str, tuple[str, str]] = {}
    for metric in plan.metrics:
        if metric.aggregation == "count":
            agg_map[metric.name] = (metric.source_column, "count")
        elif metric.aggregation == "avg":
            agg_map[metric.name] = (metric.source_column, "mean")
        else:
            agg_map[metric.name] = (metric.source_column, metric.aggregation)

    grouped = frame.groupby(plan.group_by, dropna=False).agg(**agg_map).reset_index()
    if plan.name in workbook.sheetnames:
        del workbook[plan.name]
    summary_sheet = workbook.create_sheet(title=plan.name)
    summary_sheet.append(list(grouped.columns))
    for row in grouped.itertuples(index=False):
        summary_sheet.append(list(row))
    apply_style_options(summary_sheet, plan.style)


def _format_existing_sheet(workbook, plan: SheetPlan) -> None:
    target_name = plan.source_sheet or plan.name
    if target_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{target_name}' not found for formatting.")
    apply_style_options(workbook[target_name], plan.style)


def _clean_sheet(sheet, plan: SheetPlan) -> None:
    if not plan.clean:
        return

    if plan.clean.trim_text:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if isinstance(cell.value, str):
                    cell.value = cell.value.strip()

    if plan.clean.remove_empty_rows:
        rows_to_delete: list[int] = []
        start_row = plan.data_start_row or (plan.header_row + 1 if plan.header_row else 2)
        for row_idx in range(start_row, sheet.max_row + 1):
            values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
            if all(value in (None, "") for value in values):
                rows_to_delete.append(row_idx)
        for row_idx in reversed(rows_to_delete):
            sheet.delete_rows(row_idx, 1)


def _coerce_sort_value(value, numeric: bool):
    if value is None:
        return float("-inf") if numeric else ""
    if not numeric:
        return str(value)

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        normalized = value.replace("¥", "").replace("元", "").replace(",", "").strip()
        try:
            return float(normalized)
        except ValueError:
            return float("-inf")

    return float("-inf")


def _resolve_sort_column_index(sheet, plan: SheetPlan) -> int:
    if not plan.sort:
        raise ValueError("Sort operation requires sort configuration.")
    header_row = plan.header_row or 1
    for col_idx in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=header_row, column=col_idx).value
        if header_value is not None and str(header_value).strip() == plan.sort.column:
            return col_idx
    raise ValueError(f"Sort column '{plan.sort.column}' not found in header row {header_row}.")


def _data_area_has_merged_cells(sheet, data_start_row: int, data_end_row: int) -> bool:
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.max_row >= data_start_row and merged_range.min_row <= data_end_row:
            return True
    return False


def _sort_sheet(sheet, plan: SheetPlan) -> None:
    if not plan.sort:
        raise ValueError("sort_rows operation requires sort configuration.")

    header_row = plan.header_row or 1
    data_start_row = plan.data_start_row or (header_row + 1)
    data_end_row = plan.data_end_row or sheet.max_row
    if data_start_row > data_end_row:
        return

    if _data_area_has_merged_cells(sheet, data_start_row, data_end_row):
        raise ValueError("数据区域包含合并单元格，无法安全排序。")

    sort_col_idx = _resolve_sort_column_index(sheet, plan)
    rows = []
    for row_idx in range(data_start_row, data_end_row + 1):
        row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
        rows.append(row_values)

    rows.sort(
        key=lambda row: _coerce_sort_value(row[sort_col_idx - 1], plan.sort.numeric),
        reverse=plan.sort.order == "desc",
    )

    for offset, row_values in enumerate(rows, start=data_start_row):
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row=offset, column=col_idx, value=value)


def resolve_target_date_columns(sheet, plan: SheetPlan) -> list[int]:
    header_row = plan.header_row or 1
    headers = [
        str(sheet.cell(row=header_row, column=col_idx).value).strip()
        if sheet.cell(row=header_row, column=col_idx).value is not None
        else ""
        for col_idx in range(1, sheet.max_column + 1)
    ]
    requested = plan.date_update.target_columns if plan.date_update else []
    if requested:
        return [
            idx
            for idx, header in enumerate(headers, start=1)
            if header and header in requested
        ]
    return [
        idx
        for idx, header in enumerate(headers, start=1)
        if header and "日期" in header
    ]


def _update_date_month(sheet, plan: SheetPlan) -> dict[str, object]:
    if not plan.date_update:
        raise ValueError("update_date_month requires date_update.")
    target_columns = resolve_target_date_columns(sheet, plan)
    if not target_columns:
        raise ValueError("未找到日期列，请确认表头中包含“日期”字段。")

    data_start_row = plan.data_start_row or ((plan.header_row or 1) + 1)
    updated_cells = 0
    for row_idx in range(data_start_row, sheet.max_row + 1):
        for col_idx in target_columns:
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = cell.value
            if isinstance(value, datetime):
                cell.value = value.replace(month=plan.date_update.target_month)
                updated_cells += 1
                continue
            if isinstance(value, str):
                normalized = value.strip().replace("/", "-").replace(".", "-")
                try:
                    parsed = datetime.fromisoformat(normalized)
                except ValueError:
                    continue
                cell.value = parsed.replace(month=plan.date_update.target_month)
                updated_cells += 1
    return {"updated_cells": updated_cells, "target_columns": target_columns}


def _fill_column_with_value(sheet, plan: SheetPlan) -> dict[str, object]:
    if not plan.fill:
        raise ValueError("fill_column_with_value requires fill.")

    fill_plan = plan.fill
    header_row = plan.header_row or 1
    column_index = _ensure_column(sheet, fill_plan.column_name, header_row)
    data_start_row = plan.data_start_row or ((plan.header_row or 1) + 1)
    updated_cells = 0

    if fill_plan.value_mode == "today_datetime":
        fill_value = datetime.now()
        number_format = "yyyy-mm-dd hh:mm:ss"
    elif fill_plan.value_mode == "today_date":
        fill_value = datetime.now().date()
        number_format = "yyyy-mm-dd"
    else:
        fill_value = fill_plan.static_value or ""
        number_format = None

    target_last_row = max(sheet.max_row, data_start_row)
    for row_idx in range(data_start_row, target_last_row + 1):
        cell = sheet.cell(row=row_idx, column=column_index)
        if cell.value not in (None, "") and not fill_plan.overwrite_existing:
            continue
        cell.value = fill_value
        if number_format:
            cell.number_format = number_format
        updated_cells += 1

    return {
        "updated_cells": updated_cells,
        "column_name": fill_plan.column_name,
        "column_index": column_index,
    }


def modify_workbook_from_plan(
    plan: ExcelPlan,
    uploaded_path: str | Path,
    output_path: str | Path,
) -> Path:
    shutil.copy2(uploaded_path, output_path)
    workbook = load_workbook_safe(output_path, data_only=False)

    for sheet_plan in plan.sheets:
        if sheet_plan.operation == "append_columns":
            target_name = sheet_plan.source_sheet or sheet_plan.name
            if target_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{target_name}' not found for append_columns.")
            _append_columns(workbook[target_name], sheet_plan.columns or [], sheet_plan.header_row or 1)
            apply_style_options(workbook[target_name], sheet_plan.style)
        elif sheet_plan.operation == "create_summary_sheet":
            _create_summary_from_source(workbook, sheet_plan)
        elif sheet_plan.operation == "format_sheet":
            _format_existing_sheet(workbook, sheet_plan)
        elif sheet_plan.operation == "clean_sheet":
            target_name = sheet_plan.source_sheet or sheet_plan.name
            if target_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{target_name}' not found for clean_sheet.")
            _clean_sheet(workbook[target_name], sheet_plan)
        elif sheet_plan.operation == "sort_rows":
            target_name = sheet_plan.source_sheet or sheet_plan.name
            if target_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{target_name}' not found for sort_rows.")
            _sort_sheet(workbook[target_name], sheet_plan)
        elif sheet_plan.operation == "format_and_sort_sheet":
            target_name = sheet_plan.source_sheet or sheet_plan.name
            if target_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{target_name}' not found for format_and_sort_sheet.")
            sheet = workbook[target_name]
            _clean_sheet(sheet, sheet_plan)
            _sort_sheet(sheet, sheet_plan)
            apply_style_options(sheet, sheet_plan.style)
        elif sheet_plan.operation == "create_sheet":
            if sheet_plan.name in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_plan.name}' already exists.")
            sheet = workbook.create_sheet(title=sheet_plan.name)
            if sheet_plan.columns:
                sheet.append(sheet_plan.columns)
            apply_style_options(sheet, sheet_plan.style)
        elif sheet_plan.operation == "update_date_month":
            target_name = sheet_plan.source_sheet or sheet_plan.name
            if target_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{target_name}' not found for update_date_month.")
            _update_date_month(workbook[target_name], sheet_plan)
        elif sheet_plan.operation == "fill_column_with_value":
            target_name = sheet_plan.source_sheet or sheet_plan.name
            if target_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{target_name}' not found for fill_column_with_value.")
            _fill_column_with_value(workbook[target_name], sheet_plan)
        else:
            raise ValueError(f"Unsupported operation: {sheet_plan.operation}")

    workbook.save(output_path)
    return Path(output_path)


def copy_workbook_for_simple_execution(uploaded_path: str | Path, output_path: str | Path):
    shutil.copy2(uploaded_path, output_path)
    return load_workbook_safe(output_path, data_only=False)


def save_simple_execution_workbook(workbook, output_path: str | Path) -> Path:
    workbook.save(output_path)
    return Path(output_path)


def apply_named_style_step(sheet, action_name: str) -> None:
    if action_name == "freeze_header":
        apply_freeze_header(sheet)
        return
    if action_name == "auto_filter":
        apply_auto_filter(sheet)
        return
    if action_name == "auto_width":
        apply_auto_width(sheet)
        return
    if action_name == "header_bold":
        apply_header_bold(sheet)
        return
    raise ValueError(f"Unsupported style action: {action_name}")


def execute_simple_sheet_operation(workbook, sheet_plan: SheetPlan) -> None:
    if sheet_plan.operation == "append_columns":
        target_name = sheet_plan.source_sheet or sheet_plan.name
        if target_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_name}' not found for append_columns.")
        _append_columns(workbook[target_name], sheet_plan.columns or [], sheet_plan.header_row or 1)
        return

    if sheet_plan.operation == "create_summary_sheet":
        _create_summary_from_source(workbook, sheet_plan)
        return

    if sheet_plan.operation == "format_sheet":
        target_name = sheet_plan.source_sheet or sheet_plan.name
        if target_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_name}' not found for format_sheet.")
        return

    if sheet_plan.operation == "clean_sheet":
        target_name = sheet_plan.source_sheet or sheet_plan.name
        if target_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_name}' not found for clean_sheet.")
        _clean_sheet(workbook[target_name], sheet_plan)
        return

    if sheet_plan.operation == "sort_rows":
        target_name = sheet_plan.source_sheet or sheet_plan.name
        if target_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_name}' not found for sort_rows.")
        _sort_sheet(workbook[target_name], sheet_plan)
        return

    if sheet_plan.operation == "format_and_sort_sheet":
        target_name = sheet_plan.source_sheet or sheet_plan.name
        if target_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_name}' not found for format_and_sort_sheet.")
        sheet = workbook[target_name]
        _clean_sheet(sheet, sheet_plan)
        _sort_sheet(sheet, sheet_plan)
        return

    if sheet_plan.operation == "create_sheet":
        if sheet_plan.name in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_plan.name}' already exists.")
        sheet = workbook.create_sheet(title=sheet_plan.name)
        if sheet_plan.columns:
            sheet.append(sheet_plan.columns)
        return

    if sheet_plan.operation == "update_date_month":
        target_name = sheet_plan.source_sheet or sheet_plan.name
        if target_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_name}' not found for update_date_month.")
        _update_date_month(workbook[target_name], sheet_plan)
        return

    if sheet_plan.operation == "fill_column_with_value":
        target_name = sheet_plan.source_sheet or sheet_plan.name
        if target_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{target_name}' not found for fill_column_with_value.")
        _fill_column_with_value(workbook[target_name], sheet_plan)
        return

    raise ValueError(f"Unsupported operation: {sheet_plan.operation}")
