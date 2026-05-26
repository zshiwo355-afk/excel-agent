from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.excel_tools.formatter import apply_style_options
from app.excel_tools.reader import load_workbook_safe
from app.schemas.excel_plan import ExcelPlan


def _normalize_header(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _sheet_context_for(workbook_contexts: list[dict[str, Any]], file_id: str, sheet_name: str) -> dict[str, Any]:
    for workbook in workbook_contexts:
        if workbook.get("file_id") != file_id:
            continue
        for sheet in workbook.get("sheets", []):
            if sheet.get("name") == sheet_name:
                return sheet
    raise ValueError(f"Missing workbook context for {file_id}/{sheet_name}")


def _output_headers(plan: ExcelPlan, workbook_contexts: list[dict[str, Any]]) -> list[str]:
    if plan.merge is None:
        return []
    if plan.merge.column_mapping:
        headers = list(plan.merge.column_mapping.keys())
    else:
        headers = []
        for source_sheet in plan.merge.source_sheets:
            sheet_context = _sheet_context_for(workbook_contexts, source_sheet.file_id, source_sheet.sheet_name)
            for header in sheet_context.get("headers", []):
                normalized = _normalize_header(header)
                if normalized and normalized not in headers:
                    headers.append(normalized)
    if plan.merge.add_source_columns:
        for source_column in plan.merge.source_columns:
            if source_column not in headers:
                headers.append(source_column)
    return headers


def _resolve_source_header(target_header: str, source_headers: list[str], column_mapping: dict[str, list[str]]) -> str | None:
    if target_header in column_mapping:
        for alias in column_mapping[target_header]:
            if alias in source_headers:
                return alias
    return target_header if target_header in source_headers else None


def merge_workbooks_by_plan(
    plan: ExcelPlan,
    workbook_contexts: list[dict[str, Any]],
    output_path: str | Path,
) -> dict[str, Any]:
    if not plan.merge:
        raise ValueError("merge_workbooks plan requires merge.")

    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = plan.merge.target_sheet_name
    headers = _output_headers(plan, workbook_contexts)
    detail_sheet.append(headers)

    summary_sheet = workbook.create_sheet(title="合并说明")
    summary_sheet.append(["来源文件", "来源Sheet", "原始行数", "合并行数"])

    file_index = {item.get("file_id"): item for item in workbook_contexts}
    merged_rows = 0
    merge_summary: list[dict[str, Any]] = []

    for source in plan.merge.source_sheets:
        workbook_context = file_index.get(source.file_id)
        if not workbook_context:
            raise ValueError(f"Workbook context not found for file_id={source.file_id}")
        workbook_path = workbook_context.get("file_path")
        source_workbook = load_workbook_safe(workbook_path, data_only=True)
        if source.sheet_name not in source_workbook.sheetnames:
            raise ValueError(f"Sheet '{source.sheet_name}' not found in {workbook_context.get('file_name')}")
        source_sheet = source_workbook[source.sheet_name]
        header_values = [
            _normalize_header(source_sheet.cell(row=source.header_row, column=col_idx).value)
            for col_idx in range(1, source_sheet.max_column + 1)
        ]
        source_headers = [header for header in header_values if header]
        original_rows = 0
        appended_rows = 0

        for row_idx in range(source.data_start_row, source_sheet.max_row + 1):
            row_values = [
                source_sheet.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, source_sheet.max_column + 1)
            ]
            if all(value in (None, "") for value in row_values):
                continue
            original_rows += 1
            row_dict = {
                source_headers[idx]: row_values[idx]
                for idx in range(min(len(source_headers), len(row_values)))
                if source_headers[idx]
            }
            output_row: list[Any] = []
            for header in headers:
                if plan.merge.add_source_columns and header == plan.merge.source_columns[0]:
                    output_row.append(source.file_name)
                    continue
                if plan.merge.add_source_columns and len(plan.merge.source_columns) > 1 and header == plan.merge.source_columns[1]:
                    output_row.append(source.sheet_name)
                    continue
                mapped_header = _resolve_source_header(header, source_headers, plan.merge.column_mapping)
                output_row.append(row_dict.get(mapped_header))
            detail_sheet.append(output_row)
            appended_rows += 1
            merged_rows += 1

        summary_sheet.append([source.file_name, source.sheet_name, original_rows, appended_rows])
        merge_summary.append(
            {
                "file_name": source.file_name,
                "sheet_name": source.sheet_name,
                "original_rows": original_rows,
                "merged_rows": appended_rows,
            }
        )

    apply_style_options(detail_sheet, plan.style)
    apply_style_options(summary_sheet, plan.style)
    workbook.save(output_path)

    return {
        "merged_rows": merged_rows,
        "headers": headers,
        "summary": merge_summary,
    }
