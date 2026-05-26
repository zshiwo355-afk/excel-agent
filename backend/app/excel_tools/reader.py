from pathlib import Path

from openpyxl import load_workbook


def load_workbook_safe(file_path: str | Path, data_only: bool = False):
    return load_workbook(filename=file_path, data_only=data_only)
