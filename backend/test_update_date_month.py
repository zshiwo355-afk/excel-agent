from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.agent.graph import execute_graph, plan_graph
from app.agent.state import AgentState


BASE_DIR = Path(__file__).resolve().parent
TMP_DIR = BASE_DIR / "storage" / "tmp_update_date_test"


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["日期", "事项"])
    sheet.append([datetime(2026, 5, 8), "A"])
    sheet.append([datetime(2026, 5, 18), "B"])
    workbook.save(path)


def run_update_date_month_test() -> None:
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    file_path = TMP_DIR / "input.xlsx"
    write_workbook(file_path)

    state = AgentState(
        task_id="update-date-test",
        message="帮我把日期修改为7月",
        uploaded_file_path=str(file_path),
        uploaded_file_paths=[str(file_path)],
        uploaded_files=[
            {
                "file_id": "file_1",
                "file_name": "input.xlsx",
                "file_path": str(file_path),
                "size": file_path.stat().st_size,
            }
        ],
        logs=["Task created."],
        status="planning",
    )

    planned = AgentState.model_validate(plan_graph.invoke(state))
    assert planned.excel_plan is not None
    assert planned.excel_plan["sheets"][0]["operation"] == "update_date_month"

    executed = AgentState.model_validate(execute_graph.invoke(planned.model_copy(update={"status": "running"})))
    assert executed.status == "completed"
    assert executed.output_file_path

    titles = [step.title for step in executed.execution_steps]
    assert "识别日期列" in titles
    assert "修改日期为 7 月" in titles
    assert "生成结果文件" in titles
    assert "校验结果文件" in titles

    workbook = load_workbook(executed.output_file_path)
    sheet = workbook["Sheet1"]
    assert sheet["A2"].value.month == 7
    assert sheet["A3"].value.month == 7
    assert sheet["A2"].value.day == 8
    assert sheet["A3"].value.day == 18


if __name__ == "__main__":
    run_update_date_month_test()
    print("update date month test passed")
